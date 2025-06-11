from PyQt6.QtWidgets import QApplication, QLabel, QWidget, QPushButton, QListWidget, QListWidgetItem, QLineEdit, QTableWidget, QTableWidgetItem, QHeaderView, QComboBox
from PyQt6.QtCore import Qt, QSize, QPropertyAnimation, QRect, QEasingCurve
from PyQt6.QtGui import QIcon
from datetime import datetime as dt, timedelta
import openpyxl.styles
import win32com.client as client
from tkinter import filedialog
import polars as pl
import numpy as np
import pythoncom
import openpyxl
import chardet
import json
import os
import re

path = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/")+"/"

# key = Monat in dem geprüft wird, value = Zeitspanne der Prüfung
intervall_map = {
    5: {"start": 1, "end": 4},
    8: {"start": 5, "end": 7},
    11: {"start": 8, "end": 10},
    2: {"start": 11, "end": 1}
}

month_map = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember"
}

# QPushButton, der bei klick seinen Text kopiert
class Copy_Button(QPushButton):
    def __init__(self, p_body, p_path, content):
        super().__init__(p_body)
        self.content = content
        self.p_path = p_path
        self.setText(content)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clicked.connect(self.copy_to_clipboard)

        self.copy_label = QPushButton(self)
        self.copy_label.setIcon(QIcon(p_path + "copy-outline.svg"))
        self.copy_label.setIconSize(QSize(15, 15))
        self.copy_label.setStyleSheet("text-align: center; background: transparent; border: none;")

        self.setStyleSheet("""
                           QPushButton {
                           text-align: left;
                           padding-left: 5px;
                           background: rgba(50, 50, 200, 0.1);
                           border: none;
                           }
                           QPushButton:hover {
                           background-color: rgba(50, 50, 200, 0.2);
                           }
                           """)

    def copy_to_clipboard(self):
        """Copy the Buttons Text to the clipboard"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.content)
        self.copy_label.setIcon(QIcon(self.p_path + "checkmark-outline.svg"))

    def uncheck(self):
        """Set the Buttons Icon back to default"""
        self.copy_label.setIcon(QIcon(self.p_path + "copy-outline.svg"))

    def get_body(self):
        return self

    def get_copy_label(self):
        return self.copy_label

# Kombination aus QLabel (Nameskürzel und Mantisnummer) und Textfeld (QLineEdit) für E-Mail-Adresse
class Recipient(QLabel):
    def __init__(self, name: str, email: str):
        super().__init__()
        self.name = name
        self.email = email

        self.body = QLabel(self)
        self.name_label = QLabel(self.body)
        self.name_label.setText(name)

        self.email_entry = QLineEdit(self.body)
        self.email_entry.setPlaceholderText("E-Mail Adresse eingeben")
        if len(self.email) > 0:
            self.email_entry.setText(self.email)

        self.body.setObjectName("rec_body")
        self.name_label.setObjectName("rec_name")
        self.email_entry.setObjectName("rec_entry")

    def show(self):
        self.body.show()

    def hide(self):
        self.body.hide()

# Funktionalitäten für Window
class Functions:
    def __init__(self, parent):
        self.parent = parent
        self.recipients_map = {}
        self.recipients_mantis_map = {}
        self.start = 0
        self.end = 0
        self.bold_font = openpyxl.styles.Font(bold=True)
        self.gray_fill = openpyxl.styles.PatternFill(start_color="A0A0A0", end_color="A0A0A0", fill_type="solid")
        self.yellow_fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

        # Umgebungsvariable USERNAME
        try:
            self.user_name = os.environ["USERNAME"]
        except:
            self.user_name = "User"


    def load_file(self):
        """Load Excel files Sheetnames and creates all DataFrames"""
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self.file_name = path.replace("\\", "/").split("/")[-1]
        self.parent.path = path
        self.sheet_names = openpyxl.load_workbook(path).sheetnames
        self.sheet_names = sorted(self.sheet_names)
        self.parent.sheet_names = self.sheet_names
        self.collect_dataframes()


    def df_from_sheet(self, sheet_name):
        """Convert openpyxl Sheet into polars DataFrame"""
        df = pl.read_excel(source=self.parent.path, sheet_name=sheet_name)
        extra_cols = pl.DataFrame({
            "Stichprobe": [f"{'_'.join(str(sheet_name).split(' ')[:2])}_SP{i+1:02d}" for i in range(df.shape[0])],
            "Stichprobe_": [" " for i in range(df.shape[0])]
        })

        df = extra_cols.with_columns([df[col] for col in df.columns])
        for c in df.columns:
            if df[c].dtype != pl.String:
                df = df.with_columns(df[c].cast(pl.String).alias(c))

        num_picks = 2
        if "Sammler" in sheet_name:
            num_picks = 3

        picks = np.random.choice(range(df.shape[0]), size=num_picks, replace=False)
        pick_list = pl.Series(name="Stichprobe_", values=["X" if i in picks else " " for i in range(df.shape[0])])

        df = df.with_columns(pick_list.alias("Stichprobe_"))

        return df


    def collect_dataframes(self):
        """Based on the list sheet_names, for each create a DataFrame"""
        self.full_dfs = [self.df_from_sheet(name) for name in self.parent.sheet_names]
        self.sheet_dfs = [df.filter(pl.col("Stichprobe_") == "X") for df in self.full_dfs]
        self.parent.file_import_btn.hide()
        self.parent.show_second_step_elements()
        self.parent.set_progress_perc()
        self.parent.step_name.setText("Stichprobenübersicht")


    def show_third_step(self):
        self.parent.step_number += 1
        self.parent.set_progress_perc()
        self.parent.e20_table.hide()
        self.parent.e20_collector_table.hide()
        self.parent.e30_table.hide()
        self.parent.e30_collector_table.hide()
        self.parent.redo_tables_btn.hide()
        self.parent.accept_tables_btn.hide()
        self.provide_recipients()


    def provide_recipients(self):
        """Create the list of recipients"""
        columns = []
        for df in self.sheet_dfs:
            columns.append(df["Bezeichnung"].to_list())

        # Namen des Empfängers aus der Liste nehmen (Index anhängig von singles_len)
        recipients_raw = []
        for tab in columns:
            for e in tab:
                singles = str(e).split(";")
                singles_len = len(singles)
                if singles_len == 4:
                    recipients_raw.append(singles[2])
                    self.recipients_mantis_map[singles[2].lower()] = singles[1]
                elif singles_len == 5:
                    recipients_raw.append(singles[3])
                    self.recipients_mantis_map[singles[3].lower()] = singles[2]

        # Alle einzigartigen Empfänger finden und in Kleinbuchstaben umwandeln
        self.unique_recipients = pl.Series(recipients_raw).unique().str.to_lowercase()

        # Dict mit den unique-Werten befüllen
        for r in self.unique_recipients.to_list():
            self.recipients_map[r] = ""

        #database = open("database.csv", "r")
        # dont set the json.load result equal to the dict
        database2 = json.load(open(path + "database.json"))
        #print(database2)

        # Dict durch Datenbank ergänzen (name und email)
        #for row in database.read().split(","):
        #    row_elems = row.split(":")
        #    if len(row_elems) > 1:
        #        name = row_elems[0]
        #        email = row_elems[1]
        #        self.recipients_map[name] = email
        for dbkey, dbval in database2.items():
            self.recipients_map[dbkey] = dbval
        

        #database.close()

        # Basierend auf Dict, GUI-Elemente erzeugen
        for k, v in self.recipients_map.items():
            if k in self.unique_recipients:
                item = QListWidgetItem()
                self.parent.recipient_container.addItem(item)
                rec = Recipient(f"{k} ({self.recipients_mantis_map[k]})", v)
                rec.show()
                self.parent.recipient_container.setItemWidget(item, rec)

        self.parent.set_progress_perc()
        self.parent.recipient_container.show()
        self.parent.accept_recipients_btn.show()
        self.parent.resize_recipients()
        self.parent.step_name.setText("Empfängerübersicht")


    def accept_recipients(self):
        """Get empty Recipients Fields and update the recipients map. Updated map gets saved to database"""
        self.missing_recipients = []
        for i in range(self.parent.recipient_container.count()):
            item = self.parent.recipient_container.item(i)
            widget = self.parent.recipient_container.itemWidget(item)
            email = widget.email_entry.text()
            if len(email.replace(" ", "")) == 0:
                self.missing_recipients.append(widget.name_label.text().rstrip(" "))

        for i in range(self.parent.recipient_container.count()):
            item = self.parent.recipient_container.item(i)
            widget = self.parent.recipient_container.itemWidget(item)
            name = widget.name_label.text().split("(")[0].rstrip(" ")
            email = widget.email_entry.text()
            if len(email) > 0:
                self.recipients_map[name] = email

        #database_str = ""
        #database = open("database.csv", "w")

        #for k, v in self.recipients_map.items():
        #    if len(v) > 0:
        #        database_str += f"{k}:{v},"

        #database.write(database_str[:-1])
        #database.close()

        to_delete_keys = []
        for k, v in self.recipients_map.items():
            if len(v) == 0:
                to_delete_keys.append(k)

        for k in to_delete_keys:
            del self.recipients_map[k]
                

        with open(path + "database.json", "w") as db_out:
            json.dump(self.recipients_map, db_out, indent=4)

        self.parent.recipient_container.hide()
        self.parent.accept_recipients_btn.hide()
        self.parent.show_fourth_step()


    def get_from_to_months(self):
        """Get start and end month of the Intervall"""
        global intervall_map
        current_month = dt.now().month # aktuellen Monat
        # Wenn aktuellen Monat in einem Prüfmonat liegt
        if current_month in list(intervall_map.keys()):
            start, end = intervall_map[current_month]["start"], intervall_map[current_month]["end"]
        # Falls nicht wird der Intervall gewählt, der den geringsten Abstand zum aktuellen Zeitpunkt hat
        else:
            dist_dict = {}
            for key in list(intervall_map.keys()):
                dist_dict[key] = key - current_month

            lowest = -100
            for k, v in dist_dict.items():
                if v <= 0 and v > lowest:
                    start = intervall_map[k]["start"]
                    end = intervall_map[k]["end"]
                    lowest = v

        self.start = start
        self.end = end

        self.parent.intervall_dropdown.setCurrentText(f"Von {start:02d} bis {end:02d}")
        self.parent.root_folder_copy_button.setText(self.date_to_dir_name())


    def df_to_html(self, df: pl.DataFrame):
        """Convert polars DataFrame to HTML string"""
        html_string = "<table style='border-collapse: collapse;'>"
        html_string += "<tr>" + "".join([f"<td style='background-color: rgb(160,160,160); border: 1px solid black; font-weight: bold; padding: 5px;'>{column}</td>" for column in df.columns]) + "</tr>"

        for i in range(df.shape[0]):
            row = df.row(index=i)
            html_string += "<tr>" + "".join([f"<td style='background-color: yellow; border: 1px solid black; padding: 5px;'>{text}</td>" for text in row]) + "</tr>"

        html_string += "</table>"
        return html_string


    def dataframe_to_worksheet(self, df: pl.DataFrame, ws):
        """Convert polars DataFrame in openpyxl worksheet"""
        ws.append(df.columns)
        for row in df.to_numpy().tolist():
            ws.append(row)


    def apply_coloring_to_worksheet(self, ws):
        for cell in ws[1]:
            cell.fill = self.gray_fill
            cell.font = self.bold_font
            cell.border = self.border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if row[1].value == "X":
                for cell in row:
                    cell.fill = self.yellow_fill
                    cell.border = self.border
            else:
                for cell in row:
                    cell.border = self.border


    def signature_as_html(self, name=None):
        """read the swb_default signature file and convert to HTML string"""
        sig_found = False

        try:
            sig_dir_path = f"C:/Users/{os.environ['USERNAME']}/AppData/Roaming/Microsoft/Signatures/"

            sig_dir = os.listdir(sig_dir_path)
            for sig in sig_dir:
                match = re.search(r"swb_default(.*)\.txt", sig)
                if match:
                    sig_file_name = match.string
                    sig_found = True

            if not sig_found and not name:
                for sig in sig_dir:
                    if sig.endswith(".txt"):
                        sig_file_name = sig

            if name:
                sig_file_name = name

            sig_file_path = sig_dir_path + sig_file_name
            sig_file = open(sig_file_path, "rb")
            sig_html = sig_file.read()
            encoding_result = chardet.detect(sig_html)["encoding"]
            sig_html = sig_html.decode(encoding_result)
            sig_html = sig_html.split("\n")

            sig_html = [e.replace("\r", "") for e in sig_html]

            html_string = ""
            for e in sig_html:
                if len(e.replace(" ", "")) == 0:
                    html_string += "<br>"
                elif e.startswith("http"):
                    html_string += f"<a href='{e}'>{e}</a>"
                else:
                    html_string += e + "<br>"
        except:
            html_string = "<h1 style='font-weight: bold; color: red;'>ERROR READING SIGNATURE</h1>"

        return html_string


    def get_from_to_month_ints(self):
        val = self.parent.intervall_dropdown.currentText()
        start, end = val.split(" bis ")
        start = start.replace("Von ", "")
        start, end = int(start), int(end)
        return start, end


    def date_to_dir_name(self):
        """Format the start and end months into Directory name"""
        self.start, self.end = self.get_from_to_month_ints()
        current_year = dt.now().year
        if self.start > self.end: # 11_2023_bis_2024_01
            root_folder_text = f"{current_year-1}_{self.start:02d}_bis_{current_year}_{self.end:02d}"
        else:
            root_folder_text = f"{current_year}_{self.start}_bis_{self.end}"
        return root_folder_text


    def set_root_dir_btn_text(self):
        """Set Text of Copy_Button instance responsible for the root directory"""
        t = self.date_to_dir_name()
        self.parent.root_folder_copy_button.content = t
        self.parent.root_folder_copy_button.setText(t)
        self.parent.root_folder_copy_button.uncheck()


    def create_email(self):
        """Create E-Mail and Excel file"""
        if len(self.parent.sharepoint_root_entry.text().replace(" ", "")) == 0:
            return

        self.parent.step_name.setText("Abschluss")
        self.parent.step_number += 1
        self.parent.set_progress_perc()

        pythoncom.CoInitialize()

        self.start, self.end = self.get_from_to_month_ints()

        outlook = client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)

        mail_html = ""

        if len(self.missing_recipients) > 0:
            mail_html += f"<span style='color: red; font-weight: bold; font-size: 24px;'>Fehlende Adresse{'n' if len(self.missing_recipients)>1 else ''} für: {', '.join(self.missing_recipients)}</span><br>"

        mail_html += f"<span style='font-family: Calibri, sans-serif;'>Liebe Kolleg*innen,<br><br>die Prüfung der Transporte vom {month_map[self.start]} bis {month_map[self.end]} steht an.<br>Hier sind die Stichproben:<br>{'<br>'.join([self.df_to_html(mail_df) for mail_df in self.sheet_dfs])}<br>Legt eure Dokumentationen bitte bis zum <span style='font-weight: bold; font-size: 20px;'>{(dt.now() + timedelta(weeks=3)).strftime('%d.%m.%Y')}</span> in die entsprechenden Unterordner ab: <a href='{self.parent.sharepoint_root_entry.text().replace(' ', '')}'>Sharepoint Ordner</a><br>{self.signature_as_html()}"
        mail.Subject = f"Transportprüfung von {self.start} bis {self.end}"
        mail.HTMLBody = mail_html

        for e in self.unique_recipients:
            if e in list(self.recipients_map.keys()):
                mail.To = ';'.join([self.recipients_map[e] for e in self.unique_recipients if e in list(self.recipients_map.keys())])

        mail.Display()
        pythoncom.CoUninitialize()

        for i, df_f in enumerate(self.full_dfs):
            df_f = df_f.with_columns(pl.Series(name=f"Anmerkungen {self.user_name}", values=["" for i in range(df_f.shape[0])]))
            self.full_dfs[i] = df_f

        excel_output_file = openpyxl.Workbook()

        e20_sheet = excel_output_file.active
        e20_sheet.title = self.sheet_names[0]
        e20_collector_sheet = excel_output_file.create_sheet(title=self.sheet_names[1])
        e30_sheet = excel_output_file.create_sheet(title=self.sheet_names[2])
        e30_collector_sheet = excel_output_file.create_sheet(title=self.sheet_names[3])

        # create individual worksheets from pl.DataFrames
        self.dataframe_to_worksheet(self.full_dfs[0], e20_sheet)
        self.dataframe_to_worksheet(self.full_dfs[1], e20_collector_sheet)
        self.dataframe_to_worksheet(self.full_dfs[2], e30_sheet)
        self.dataframe_to_worksheet(self.full_dfs[3], e30_collector_sheet)

        # filter by X in second column
        e20_sheet.auto_filter.ref = "A1:Z1000"
        e20_sheet.auto_filter.add_filter_column(1, ["X"])
        e20_collector_sheet.auto_filter.ref = "A1:Z1000"
        e20_collector_sheet.auto_filter.add_filter_column(1, ["X"])
        e30_sheet.auto_filter.ref = "A1:Z1000"
        e30_sheet.auto_filter.add_filter_column(1, ["X"])
        e30_collector_sheet.auto_filter.ref = "A1:Z1000"
        e30_collector_sheet.auto_filter.add_filter_column(1, ["X"])

        # set colors of worksheets
        self.apply_coloring_to_worksheet(e20_sheet)
        self.apply_coloring_to_worksheet(e20_collector_sheet)
        self.apply_coloring_to_worksheet(e30_sheet)
        self.apply_coloring_to_worksheet(e30_collector_sheet)

        excel_output_file.save(os.path.dirname(os.path.abspath(__file__)).replace("\\", "/") + "/output/" + self.file_name.replace(".xlsx", "_filtered.xlsx"))

        self.parent.needed_folder_names_container.setText("Folgende Ordner müssen im Sharepoint angelegt werden")

        for folder_name_coll in [df['Stichprobe'].to_list() for df in self.sheet_dfs]:
            for folder_name in folder_name_coll:
                folder_label = Copy_Button(self.parent.needed_folder_names_container, self.parent.icon_path, folder_name)

        self.parent.needed_folder_names_container.show()
        self.parent.intervall_drop_label.hide()
        self.parent.intervall_dropdown.hide()
        self.parent.sharepoint_directory_label.hide()
        self.parent.sharepoint_root_entry.hide()
        self.parent.intervall_accept_btn.hide()
        self.parent.root_folder_step_label.hide()
        self.parent.root_folder_copy_button.hide()
        self.parent.resize_folders()

# GUI
class Window(QWidget):
    def __init__(self, style_str):
        super().__init__()
        self.icon_path = os.path.dirname(os.path.abspath(__file__)).replace("\\", "/") + "/icons_colored/"
        self.bar_progress_perc: float = 0.0
        self.total_steps = 4
        self.step_number = 1
        self.banner_height = 50
        self.path = None
        self.sheet_names = None
        self.recipient_obj_list = []
        self.functions = Functions(self)
        self.style_str = style_str

        self.resize(1150, 700)
        self.setWindowTitle("Express")
        self.setStyleSheet(style_str)
        self.setMinimumSize(950, 700)
        self.setWindowIcon(QIcon(self.icon_path + "car-outline.svg"))

        self.body = QLabel(self)
        self.body.move(0, self.banner_height)
        self.body.setObjectName("body")

        self.banner = QLabel(self)
        self.banner.setObjectName("banner")

        self.progress_back = QLabel(self.banner)
        self.progress_back.setObjectName("bar_back")

        self.progress_bar = QLabel(self.banner)
        self.progress_bar.setObjectName("bar_actual")

        self.step_name = QLabel(self.banner)
        self.step_name.setText("Excel-Datei öffnen")
        self.step_name.setObjectName("step_name")

        # first step
        self.file_import_btn = QPushButton(self.body)
        self.file_import_btn.setObjectName("import_btn")
        self.file_import_btn.clicked.connect(self.functions.load_file)
        self.file_import_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.file_import_btn.setIcon(QIcon(self.icon_path + "upload-outline.svg"))
        self.file_import_btn.setIconSize(QSize(50, 50))

        # second step
        self.e20_table = QTableWidget(self.body)
        self.e20_collector_table = QTableWidget(self.body)
        self.e30_table = QTableWidget(self.body)
        self.e30_collector_table = QTableWidget(self.body)
        self.redo_tables_btn = QPushButton(self.body)
        self.accept_tables_btn = QPushButton(self.body)

        self.redo_tables_btn.clicked.connect(self.functions.collect_dataframes)
        self.accept_tables_btn.clicked.connect(self.functions.show_third_step)

        self.redo_tables_btn.setIcon(QIcon(self.icon_path + "flip-2-outline.svg"))
        self.accept_tables_btn.setIcon(QIcon(self.icon_path + "arrow-forward-outline.svg"))

        self.redo_tables_btn.setIconSize(QSize(25, 25))
        self.accept_tables_btn.setIconSize(QSize(25, 25))

        self.redo_tables_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.accept_tables_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        self.e20_table.hide()
        self.e20_collector_table.hide()
        self.e30_table.hide()
        self.e30_collector_table.hide()
        self.redo_tables_btn.hide()
        self.accept_tables_btn.hide()

        # third step
        self.recipient_container = QListWidget(self.body)
        self.recipient_container.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.recipient_container.hide()

        self.accept_recipients_btn = QPushButton(self.body)
        self.accept_recipients_btn.setIcon(QIcon(self.icon_path + "arrow-forward-outline.svg"))
        self.accept_recipients_btn.setIconSize(QSize(25, 25))
        self.accept_recipients_btn.clicked.connect(self.functions.accept_recipients)
        self.accept_recipients_btn.hide()
        self.accept_recipients_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        # fourth step
        self.intervall_dropdown = QComboBox(self.body)
        for k, v in intervall_map.items():
            self.intervall_dropdown.addItem(f"Von {v['start']:02d} bis {v['end']:02d}")
        self.intervall_dropdown.setObjectName("inter_drop")
        self.intervall_dropdown.currentTextChanged.connect(self.functions.set_root_dir_btn_text)

        self.intervall_accept_btn = QPushButton(self.body)
        self.intervall_accept_btn.setIcon(QIcon(self.icon_path + "checkmark-outline.svg"))
        self.intervall_accept_btn.setIconSize(QSize(25, 25))
        self.intervall_accept_btn.clicked.connect(self.functions.create_email)
        self.intervall_accept_btn.setCursor(Qt.CursorShape.PointingHandCursor)

        self.sharepoint_root_entry = QLineEdit(self.body)
        self.sharepoint_root_entry.setPlaceholderText("Sharepoint-Ordner Link")
        self.sharepoint_root_entry.hide()

        self.intervall_drop_label = QLabel(self.body)
        self.intervall_drop_label.setText("1. Prüfintervall")
        self.intervall_drop_label.setObjectName("sub_heading")

        self.root_folder_step_label = QLabel(self.body)
        self.root_folder_step_label.setText("2. Folgenden Ordner anlegen")
        self.root_folder_step_label.setObjectName("sub_heading")

        self.root_folder_copy_button = Copy_Button(self.body, self.icon_path, "<<>>")

        self.sharepoint_directory_label = QLabel(self.body)
        self.sharepoint_directory_label.setText("3. Sharepoint Basis Ordner")
        self.sharepoint_directory_label.setObjectName("sub_heading")

        self.intervall_dropdown.hide()
        self.intervall_accept_btn.hide()

        self.intervall_drop_label.hide()
        self.sharepoint_directory_label.hide()

        self.root_folder_step_label.hide()
        self.root_folder_copy_button.hide()

        # fifth step
        self.needed_folder_names_container = QLabel(self.body)
        self.needed_folder_names_container.hide()
        self.needed_folder_names_container.setObjectName("needed")
        self.needed_folder_names_container.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)

    # Tabellen der Stichproben
    def resize_tables(self):
        self.e20_table.setGeometry(50, 20, self.body.width()-100, 100) # per row 50px in height
        self.e20_collector_table.setGeometry(50, 140, self.body.width()-100, 150)
        self.e30_table.setGeometry(50, 310, self.body.width()-100, 100)
        self.e30_collector_table.setGeometry(50, 430, self.body.width()-100, 150)
        self.redo_tables_btn.setGeometry(self.body.width()//2 - 45, self.body.height() - 50 - self.banner_height, 40, 40)
        self.accept_tables_btn.setGeometry(self.body.width()//2 + 5, self.body.height() - 50 - self.banner_height, 40, 40)

    # Liste der Empfänger
    def resize_recipients(self):
        self.recipient_container.setGeometry(200, 100, self.body.width() - 400, self.body.height() - 250)
        for i in range(self.recipient_container.count()):
            item = self.recipient_container.item(i)
            widget = self.recipient_container.itemWidget(item)
            widget.body.resize(self.recipient_container.width() - 2, 50)
            widget.name_label.resize(248, 50)
            widget.email_entry.setGeometry(246, 0, self.recipient_container.width()-260, 50)
            item.setSizeHint(QSize(self.recipient_container.width() - 4, 60))
        self.accept_recipients_btn.setGeometry(self.body.width()//2 - 20, self.body.height() - 50 - self.banner_height, 40, 40)

    # Dropdown des Prüfintervalls usw.
    def resize_timespan(self):
        self.intervall_drop_label.setGeometry(self.body.width()//2 - 100, self.body.height()//2 - self.banner_height - 100, 200, 20)
        self.intervall_dropdown.setGeometry(self.body.width()//2 - 100 + 15, self.body.height()//2 - self.banner_height - 75, 300, 20)

        self.root_folder_step_label.setGeometry(self.body.width()//2 - 100, self.body.height()//2 - self.banner_height - 20, 250, 25)
        self.root_folder_copy_button.setGeometry(self.body.width()//2 - 100 + 15, self.body.height()//2 - self.banner_height + 10, 200, 30)
        self.root_folder_copy_button.copy_label.setGeometry(self.root_folder_copy_button.get_body().width() - 25, 5, 20, 20)

        self.sharepoint_directory_label.setGeometry(self.body.width()//2 - 100, self.body.height()//2 - self.banner_height + 60, 300, 20)
        self.sharepoint_root_entry.setGeometry(self.body.width()//2 - 100 + 15, self.body.height()//2 - self.banner_height + 85, 300, 25)
        self.intervall_accept_btn.setGeometry(self.body.width()//2 - 20, self.body.height() - 50 - self.banner_height, 40, 40)

    # Copy_Buttons der Ordnernamen
    def resize_folders(self):
        self.needed_folder_names_container.setGeometry(200, 100, self.body.width() - 400, self.body.height() - 200)
        for i, child in enumerate(self.needed_folder_names_container.children()):
            body = child.get_body()
            copy_label = child.get_copy_label()
            body.setGeometry(0, i*35 + 25, 150, 30)
            copy_label.setGeometry(body.width() - 25, 5, 20, 20)

    # Override von PyQt-Funktion, wird aufgerufen, wenn das Fenster die Grße ändert
    def resizeEvent(self, e):
        self.body.resize(self.width(), self.height())
        self.banner.resize(self.width(), self.banner_height)
        self.progress_back.setGeometry(60, 35, self.width() - 120, 10)
        self.progress_bar.setGeometry(60, 35, int(self.progress_back.width() * self.bar_progress_perc), 10)
        self.step_name.setGeometry(60, 5, self.banner.width() - 200, 30)
        self.file_import_btn.setGeometry(self.body.width()//2 - 150, self.body.height()//2 - 80 - self.banner_height, 300, 160)
        self.resize_tables()
        self.resize_recipients()
        self.resize_timespan()
        self.resize_folders()

    # Breite des Fortschrittsbalken
    def set_progress_perc(self):
        self.bar_progress_perc = self.step_number / self.total_steps
        new_width = int(self.progress_back.width() * self.bar_progress_perc)
        self.animation = QPropertyAnimation(self.progress_bar, b"geometry")
        self.animation.setDuration(500)
        self.animation.setStartValue(self.progress_bar.geometry())
        self.animation.setEndValue(QRect(60, 35, new_width, 10))
        self.animation.setEasingCurve(QEasingCurve.Type.InOutQuad)
        self.animation.start()

    # Stichproben anzeigen
    def show_second_step_elements(self):
        dataframes = self.functions.sheet_dfs

        self.e20_table.setRowCount(2)
        self.e20_collector_table.setRowCount(3)
        self.e30_table.setRowCount(2)
        self.e30_collector_table.setRowCount(3)

        self.e20_table.setColumnCount(dataframes[0].shape[1])
        self.e20_collector_table.setColumnCount(dataframes[1].shape[1])
        self.e30_table.setColumnCount(dataframes[2].shape[1])
        self.e30_collector_table.setColumnCount(dataframes[3].shape[1])

        self.e20_table.setHorizontalHeaderLabels(dataframes[0].columns)
        self.e20_collector_table.setHorizontalHeaderLabels(dataframes[1].columns)
        self.e30_table.setHorizontalHeaderLabels(dataframes[2].columns)
        self.e30_collector_table.setHorizontalHeaderLabels(dataframes[3].columns)

        self.e20_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e20_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e20_collector_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e20_collector_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e30_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e30_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e30_collector_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.e30_collector_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        #self.e20_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        #self.e20_collector_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        #self.e30_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        #self.e30_collector_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        for row in range(self.e20_table.rowCount()):
            for column in range(self.e20_table.columnCount()):
                item = QTableWidgetItem(f"{dataframes[0][row, column]}")
                self.e20_table.setItem(row, column, item)
                
        # outdated version that doesn't work if there are not enough rows in e20_collector_table
        #for row in range(self.e20_collector_table.rowCount()):
        #    for column in range(self.e20_collector_table.columnCount()):
        #        item = QTableWidgetItem(f"{dataframes[1][row, column]}")
        #        self.e20_collector_table.setItem(row, column, item)
        
        # Robustere Version die auch Funktioniert falls zu wenige rows im e20_collector_table sind
        for row in range(self.e20_collector_table.rowCount()):
            for column in range(self.e20_collector_table.columnCount()):
                if row < dataframes[1].shape[0] and column < dataframes[1].shape[1]:
                    item = QTableWidgetItem(f"{dataframes[1][row, column]}")
                else:
                    item = QTableWidgetItem("")
                self.e20_collector_table.setItem(row, column, item)

        for row in range(self.e30_table.rowCount()):
            for column in range(self.e30_table.columnCount()):
                item = QTableWidgetItem(f"{dataframes[2][row, column]}")
                self.e30_table.setItem(row, column, item)

        for row in range(self.e30_collector_table.rowCount()):
            for column in range(self.e30_collector_table.columnCount()):
                item = QTableWidgetItem(f"{dataframes[3][row, column]}")
                self.e30_collector_table.setItem(row, column, item)

        self.e20_table.show()
        self.e20_collector_table.show()
        self.e30_table.show()
        self.e30_collector_table.show()
        self.redo_tables_btn.show()
        self.accept_tables_btn.show()
        self.resize_tables()

    # Stichproben bestätigen
    def accept_tables(self):
        self.step_number += 1
        self.recipient_container.show()
        self.accept_recipients_btn.show()
        self.resize_recipients()
        self.set_progress_perc()

    # Dropdown des Prüfintervalls usw. anzeigen
    def show_fourth_step(self):
        self.step_name.setText("Detailabfrage")
        self.intervall_dropdown.show()
        self.intervall_accept_btn.show()
        self.sharepoint_root_entry.show()
        self.intervall_drop_label.show()
        self.sharepoint_directory_label.show()
        self.functions.get_from_to_months()
        self.root_folder_step_label.show()
        self.root_folder_copy_button.show()
        self.step_number += 1
        self.set_progress_perc()

# GUI starten
def execute():
    app = QApplication([])
    win = Window(open(path +"style.css", "r").read())
    win.show()
    app.exec()